from openpyxl import load_workbook


class DoExcel:
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(file_name)

    # 如果表格中只有一列数据，返回除首行的列表，如果不是则返回json格式的所有数据
    def get_all_data(self, sheet_name):
        sheet = self.wb[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        # print('{}/{} 工作簿总行数：{} 总列数：{}'.format(self.file_name, sheet_name, max_row, max_column))

        all_data = []
        
        title_list = []
        for i in range(1, max_column+1):
            title_list.append(sheet.cell(1, i).value)
        for i in range(2, max_row+1):
            row_cell_value = []
            for j in range(1, max_column+1):
                row_cell_value.append(sheet.cell(i, j).value)
            all_data.append(dict(zip(title_list, row_cell_value)))
            
        return all_data
